"""
Admin dashboard routes: login/logout, dashboard home, customers,
conversations, orders, analytics, catalog, offers, billing, settings,
printer, delivery agents. Every route except /admin/login depends on
get_current_admin, which raises 401 for missing/invalid sessions - handled
by the exception handler registered in main.py.
"""
import io
from datetime import date, datetime, timedelta

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

from admin.auth import (
    create_session_cookie_value, get_current_admin, hash_password,
    try_get_current_admin, verify_password,
)
from admin.templating import render
from catalog import store as catalog_store
from config import config
from offers import store as offers_store
from storage import store

router = APIRouter(prefix="/admin")


def _default_date_range():
    end = date.today()
    start = end - timedelta(days=30)
    return start.isoformat(), end.isoformat()


def _month_bounds(year_month: str):
    """year_month is 'YYYY-MM' - returns (first day, last day) as ISO date strings."""
    year, month = int(year_month[:4]), int(year_month[5:7])
    start = date(year, month, 1)
    end = date(year + 1, 1, 1) - timedelta(days=1) if month == 12 else date(year, month + 1, 1) - timedelta(days=1)
    return start.isoformat(), end.isoformat()


# ---- Auth ----

@router.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if try_get_current_admin(request):
        return RedirectResponse("/admin", status_code=303)
    return render(request, "login.html", error=None)


@router.post("/login")
def login_submit(request: Request, username: str = Form(...), password: str = Form(...)):
    admin = store.get_admin_by_username(username)
    if not admin or not verify_password(password, admin["password_hash"]):
        return render(request, "login.html", error="Invalid username or password.")

    response = RedirectResponse("/admin", status_code=303)
    response.set_cookie(
        config.ADMIN_SESSION_COOKIE,
        create_session_cookie_value(admin["id"]),
        max_age=config.ADMIN_SESSION_MAX_AGE,
        httponly=True,
        samesite="lax",
    )
    return response


@router.get("/logout")
def logout():
    response = RedirectResponse("/admin/login", status_code=303)
    response.delete_cookie(config.ADMIN_SESSION_COOKIE)
    return response


# ---- Dashboard home ----

@router.get("", response_class=HTMLResponse)
def dashboard_home(request: Request, admin=Depends(get_current_admin)):
    today = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()

    stats_today = store.get_stats(today, today)
    stats_month = store.get_stats(month_start, today)
    daily = store.get_daily_counts(*_last_n_days(7))
    recent_orders = store.get_orders(*_default_date_range())[:6]

    return render(
        request, "dashboard_home.html", active_page="dashboard", admin=admin,
        stats_today=stats_today, stats_month=stats_month, daily=daily,
        recent_orders=recent_orders,
    )


def _last_n_days(n: int):
    end = date.today()
    start = end - timedelta(days=n - 1)
    return start.isoformat(), end.isoformat()


# ---- Customers ----

@router.get("/customers", response_class=HTMLResponse)
def customers_page(request: Request, admin=Depends(get_current_admin), start: str = "", end: str = ""):
    default_start, default_end = _default_date_range()
    start, end = start or default_start, end or default_end
    customers = store.get_customers_summary(start, end)
    return render(
        request, "customers.html", active_page="customers", admin=admin,
        customers=customers, start=start, end=end,
    )


# ---- Conversations ----

@router.get("/conversations", response_class=HTMLResponse)
def conversations_page(request: Request, admin=Depends(get_current_admin), start: str = "", end: str = "", phone: str = ""):
    default_start, default_end = _default_date_range()
    start, end = start or default_start, end or default_end
    customers = store.get_customers_summary(start, end)
    transcript = store.get_conversation(phone, start, end) if phone else None

    # Order history and purchase-behavior stats are shown across all time,
    # not just the transcript's date filter - the point is seeing a
    # customer's buying pattern over time, not just in the current window.
    order_history_by_month = None
    orders_by_month = {}
    favorite_items = None
    behavior_stats = None
    if phone:
        order_history_by_month = store.get_customer_orders_by_month(phone)
        for month in order_history_by_month:
            month_start, month_end = _month_bounds(month["year_month"])
            orders_by_month[month["year_month"]] = store.get_customer_orders(phone, month_start, month_end)

        favorite_items = store.get_customer_favorite_items(phone, limit=5)
        all_orders = store.get_customer_orders(phone)
        placed_orders = [o for o in all_orders if o["status"] in ("confirmed", "packed", "picked_up", "delivered")]
        if placed_orders:
            total_spent = sum(o["total"] for o in placed_orders)
            avg_order_value = total_spent / len(placed_orders)
            most_recent = max(o["created_at"] for o in placed_orders)
            days_since_last = (date.today() - datetime.fromisoformat(most_recent).date()).days
            behavior_stats = {
                "total_orders": len(placed_orders),
                "total_spent": total_spent,
                "avg_order_value": avg_order_value,
                "days_since_last": days_since_last,
            }

    return render(
        request, "conversations.html", active_page="conversations", admin=admin,
        customers=customers, start=start, end=end, phone=phone, transcript=transcript,
        order_history_by_month=order_history_by_month, orders_by_month=orders_by_month,
        favorite_items=favorite_items, behavior_stats=behavior_stats,
    )


# ---- Orders ----

@router.get("/orders", response_class=HTMLResponse)
def orders_page(request: Request, admin=Depends(get_current_admin), start: str = "", end: str = "", status: str = ""):
    default_start, default_end = _default_date_range()
    start, end = start or default_start, end or default_end
    orders = store.get_orders(start, end, status or None)
    orders_with_items = [(o, store.get_order_items(o["id"])) for o in orders]
    return render(
        request, "orders.html", active_page="orders", admin=admin,
        orders_with_items=orders_with_items, start=start, end=end, status=status,
    )


# ---- Catalog (menu) ----

@router.get("/catalog", response_class=HTMLResponse)
def catalog_page(request: Request, admin=Depends(get_current_admin), q: str = ""):
    items = catalog_store.search_items(q, limit=1000) if q.strip() else catalog_store.list_items()
    last_synced = catalog_store.last_synced_at()
    return render(
        request, "catalog.html", active_page="catalog", admin=admin,
        items=items, last_synced=last_synced, q=q,
    )


@router.post("/catalog/import")
async def catalog_import(request: Request, admin=Depends(get_current_admin), file: UploadFile = File(...)):
    import os
    import tempfile

    from catalog.excel_import import import_file

    suffix = os.path.splitext(file.filename or "upload.xlsx")[1] or ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    error = None
    count = 0
    try:
        count = import_file(tmp_path, source="excel")
    except Exception as e:
        error = str(e)
    finally:
        os.remove(tmp_path)

    items = catalog_store.list_items()
    last_synced = catalog_store.last_synced_at()
    return render(
        request, "catalog.html", active_page="catalog", admin=admin,
        items=items, last_synced=last_synced, q="",
        import_message=f"Imported/updated {count} item(s)." if not error else None,
        import_error=error,
    )


@router.post("/catalog/{item_id}/toggle-stock")
def catalog_toggle_stock(item_id: int, admin=Depends(get_current_admin), q: str = Form("")):
    item = catalog_store.get_item(item_id)
    if item:
        catalog_store.set_stock(item_id, 0 if item["in_stock"] else 1)
    return RedirectResponse(f"/admin/catalog?q={q}", status_code=303)


# ---- Analytics ----

@router.get("/analytics", response_class=HTMLResponse)
def analytics_page(request: Request, admin=Depends(get_current_admin), start: str = "", end: str = "", preset: str = ""):
    start, end = _resolve_range(start, end, preset)
    stats = store.get_stats(start, end)
    daily = store.get_daily_counts(start, end)
    orders = store.get_orders(start, end)
    return render(
        request, "analytics.html", active_page="analytics", admin=admin,
        stats=stats, daily=daily, orders=orders, start=start, end=end, preset=preset,
    )


def _resolve_range(start: str, end: str, preset: str):
    today = date.today()
    if preset == "today":
        return today.isoformat(), today.isoformat()
    if preset == "month":
        return today.replace(day=1).isoformat(), today.isoformat()
    if preset == "7d":
        return (today - timedelta(days=6)).isoformat(), today.isoformat()
    default_start, default_end = _default_date_range()
    return start or default_start, end or default_end


@router.get("/analytics/export")
def analytics_export(admin=Depends(get_current_admin), start: str = "", end: str = ""):
    default_start, default_end = _default_date_range()
    start, end = start or default_start, end or default_end

    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = "Messages"
    ws.append(["Timestamp (UTC)", "Phone", "Name", "Direction", "Message"])
    for row in store.get_all_messages(start, end):
        ws.append([row["created_at"], row["phone"], row["name"], "Customer" if row["direction"] == "in" else "Bot", row["message"]])
    for col_letter, width in zip("ABCDE", [26, 16, 20, 10, 60]):
        ws.column_dimensions[col_letter].width = width

    ws2 = wb.create_sheet("Customers")
    ws2.append(["Phone", "Name", "Message Count", "Last Message (UTC)"])
    for row in store.get_customers_summary(start, end):
        ws2.append([row["phone"], row["name"], row["message_count"], row["last_message_at"]])
    for col_letter, width in zip("ABCD", [16, 20, 14, 26]):
        ws2.column_dimensions[col_letter].width = width

    ws3 = wb.create_sheet("Daily Summary")
    ws3.append(["Date", "Messages Received", "Replies Sent"])
    for row in store.get_daily_counts(start, end):
        ws3.append([row["day"], row["received"], row["sent"]])
    for col_letter, width in zip("ABC", [14, 18, 14]):
        ws3.column_dimensions[col_letter].width = width

    ws4 = wb.create_sheet("Orders")
    ws4.append(["Order ID", "Phone", "Status", "Pickup", "Delivery Agent", "Subtotal", "Delivery Fee", "Discount Offer ID", "Total", "Created (UTC)", "Confirmed (UTC)"])
    for o in store.get_orders(start, end):
        ws4.append([o["id"], o["phone"], o["status"], "Yes" if o["is_pickup"] else "No", o["delivery_agent_phone"], o["subtotal"], o["delivery_fee"], o["discount_applied"], o["total"], o["created_at"], o["confirmed_at"]])
    for col_letter, width in zip("ABCDEFGHIJK", [10, 16, 14, 8, 16, 12, 12, 14, 12, 26, 26]):
        ws4.column_dimensions[col_letter].width = width

    ws5 = wb.create_sheet("Menu Snapshot")
    ws5.append(["SKU", "Name", "Category", "Unit", "Price", "In Stock", "Source", "Updated (UTC)"])
    for item in catalog_store.list_items():
        ws5.append([item["sku"], item["name"], item["category"], item["unit"], item["price"], "Yes" if item["in_stock"] else "No", item["source"], item["updated_at"]])
    for col_letter, width in zip("ABCDEFGH", [14, 32, 22, 12, 10, 10, 10, 26]):
        ws5.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"finedine-report_{start}_to_{end}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---- Reports ----
# Distinct from Analytics: pre-built, calendar-period summaries (monthly /
# yearly rollups, top-selling items) rather than an arbitrary custom range.

@router.get("/reports", response_class=HTMLResponse)
def reports_page(request: Request, admin=Depends(get_current_admin), year: str = "", month: str = ""):
    today = date.today()
    year = year or str(today.year)
    month = month or f"{today.month:02d}"
    month_start, month_end = _month_bounds(f"{year}-{month}")

    monthly_stats = store.get_stats(month_start, month_end)
    top_items = store.get_top_selling_items(month_start, month_end, limit=10)
    yearly_rollup = store.get_orders_by_month(f"{year}-01-01", f"{year}-12-31")

    available_years = sorted({row["year_month"][:4] for row in store.get_orders_by_month()}, reverse=True)
    if not available_years:
        available_years = [str(today.year)]

    return render(
        request, "reports.html", active_page="reports", admin=admin,
        year=year, month=month, monthly_stats=monthly_stats, top_items=top_items,
        yearly_rollup=yearly_rollup, available_years=available_years,
    )


@router.get("/reports/export")
def reports_export(admin=Depends(get_current_admin), year: str = "", month: str = ""):
    today = date.today()
    year = year or str(today.year)
    month = month or f"{today.month:02d}"
    month_start, month_end = _month_bounds(f"{year}-{month}")

    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = "Monthly Summary"
    stats = store.get_stats(month_start, month_end)
    ws.append(["Metric", "Value"])
    ws.append(["Period", f"{year}-{month}"])
    ws.append(["Messages received", stats["messages_received"]])
    ws.append(["Replies sent", stats["replies_sent"]])
    ws.append(["Unique customers", stats["unique_customers"]])
    ws.append(["Orders confirmed", stats["orders_confirmed"]])
    ws.append(["Revenue", stats["revenue"]])
    for col_letter, width in zip("AB", [24, 20]):
        ws.column_dimensions[col_letter].width = width

    ws2 = wb.create_sheet("Top Items")
    ws2.append(["Item", "Qty Sold", "Revenue"])
    for row in store.get_top_selling_items(month_start, month_end, limit=50):
        ws2.append([row["name"], row["total_qty"], row["total_revenue"]])
    for col_letter, width in zip("ABC", [30, 12, 14]):
        ws2.column_dimensions[col_letter].width = width

    ws3 = wb.create_sheet("Yearly Rollup")
    ws3.append(["Month", "Orders", "Revenue", "Unique Customers"])
    for row in store.get_orders_by_month(f"{year}-01-01", f"{year}-12-31"):
        ws3.append([row["year_month"], row["order_count"], row["total_revenue"], row["unique_customers"]])
    for col_letter, width in zip("ABCD", [12, 12, 14, 16]):
        ws3.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"finedine-report_{year}-{month}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---- Offers ----

@router.get("/offers", response_class=HTMLResponse)
def offers_page(request: Request, admin=Depends(get_current_admin)):
    offers = offers_store.list_offers()
    return render(request, "offers.html", active_page="offers", admin=admin, offers=offers, draft=None)


@router.post("/offers/draft")
def offers_draft(request: Request, admin=Depends(get_current_admin), prompt: str = Form(...)):
    from ai.offer_assistant import draft_offer

    offers = offers_store.list_offers()
    error = None
    draft = None
    try:
        draft = draft_offer(prompt)
    except Exception as e:
        error = str(e)

    return render(
        request, "offers.html", active_page="offers", admin=admin,
        offers=offers, draft=draft, draft_error=error, last_prompt=prompt,
    )


@router.post("/offers/create")
def offers_create(
    request: Request, admin=Depends(get_current_admin),
    title: str = Form(...), description: str = Form(""), discount_type: str = Form(...),
    discount_value: float = Form(...), starts_at: str = Form(""), ends_at: str = Form(""),
):
    offers_store.create_offer(
        title=title, description=description, discount_type=discount_type,
        discount_value=discount_value, starts_at=starts_at or None, ends_at=ends_at or None,
        status="draft", created_by_admin_id=admin["id"],
    )
    return RedirectResponse("/admin/offers", status_code=303)


@router.post("/offers/{offer_id}/activate")
def offers_activate(offer_id: int, admin=Depends(get_current_admin)):
    offers_store.set_offer_status(offer_id, "active")
    return RedirectResponse("/admin/offers", status_code=303)


@router.post("/offers/{offer_id}/deactivate")
def offers_deactivate(offer_id: int, admin=Depends(get_current_admin)):
    offers_store.set_offer_status(offer_id, "expired")
    return RedirectResponse("/admin/offers", status_code=303)


@router.post("/offers/{offer_id}/delete")
def offers_delete(offer_id: int, admin=Depends(get_current_admin)):
    offers_store.delete_offer(offer_id)
    return RedirectResponse("/admin/offers", status_code=303)


# ---- Delivery agents ----

@router.get("/delivery-agents", response_class=HTMLResponse)
def delivery_agents_page(request: Request, admin=Depends(get_current_admin)):
    agents = store.list_delivery_agents()
    return render(request, "delivery_agents.html", active_page="delivery_agents", admin=admin, agents=agents, error=None)


@router.post("/delivery-agents")
def delivery_agents_add(request: Request, admin=Depends(get_current_admin), phone: str = Form(...), name: str = Form(...)):
    store.add_delivery_agent(phone.strip(), name.strip())
    return RedirectResponse("/admin/delivery-agents", status_code=303)


@router.post("/delivery-agents/{agent_id}/activate")
def delivery_agents_activate(agent_id: int, admin=Depends(get_current_admin)):
    store.set_delivery_agent_active(agent_id, True)
    return RedirectResponse("/admin/delivery-agents", status_code=303)


@router.post("/delivery-agents/{agent_id}/deactivate")
def delivery_agents_deactivate(agent_id: int, admin=Depends(get_current_admin)):
    store.set_delivery_agent_active(agent_id, False)
    return RedirectResponse("/admin/delivery-agents", status_code=303)


# ---- Billing ----

@router.get("/billing", response_class=HTMLResponse)
def billing_page(request: Request, admin=Depends(get_current_admin)):
    today = date.today()
    stats_month = store.get_stats(today.replace(day=1).isoformat(), today.isoformat())
    return render(request, "billing.html", active_page="billing", admin=admin, stats_month=stats_month)


# ---- Printer ----

@router.get("/printer", response_class=HTMLResponse)
def printer_page(request: Request, admin=Depends(get_current_admin)):
    settings = store.get_printer_settings()
    return render(request, "printer.html", active_page="printer", admin=admin, settings=settings, message=None, error=None)


@router.post("/printer")
def printer_save(
    request: Request, admin=Depends(get_current_admin),
    connection_type: str = Form(...), printer_ip: str = Form(""), printer_port: str = Form(""),
):
    if connection_type not in ("network", "bluetooth", "usb"):
        return render(
            request, "printer.html", active_page="printer", admin=admin,
            settings=store.get_printer_settings(), message=None, error="Invalid connection type.",
        )
    if connection_type == "network" and not printer_ip.strip():
        return render(
            request, "printer.html", active_page="printer", admin=admin,
            settings=store.get_printer_settings(), message=None, error="Printer IP address is required for a network printer.",
        )
    if connection_type in ("bluetooth", "usb") and not printer_port.strip():
        return render(
            request, "printer.html", active_page="printer", admin=admin,
            settings=store.get_printer_settings(), message=None,
            error="COM port / device path is required for a Bluetooth or USB printer.",
        )

    store.set_printer_settings(
        connection_type=connection_type,
        printer_ip=printer_ip.strip() or None,
        printer_port=printer_port.strip() or None,
    )
    return render(
        request, "printer.html", active_page="printer", admin=admin,
        settings=store.get_printer_settings(), message="Printer settings saved.", error=None,
    )


# ---- Settings ----

@router.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request, admin=Depends(get_current_admin)):
    return render(request, "settings.html", active_page="settings", admin=admin, message=None, error=None)


@router.post("/settings/password")
def settings_change_password(
    request: Request, admin=Depends(get_current_admin),
    current_password: str = Form(...), new_password: str = Form(...),
):
    if not verify_password(current_password, admin["password_hash"]):
        return render(request, "settings.html", active_page="settings", admin=admin, message=None, error="Current password is incorrect.")
    if len(new_password) < 8:
        return render(request, "settings.html", active_page="settings", admin=admin, message=None, error="New password must be at least 8 characters.")

    store.set_admin_password(admin["id"], hash_password(new_password))
    return render(request, "settings.html", active_page="settings", admin=admin, message="Password updated.", error=None)
