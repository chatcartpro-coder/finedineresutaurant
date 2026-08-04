"""
Imports stock data from an Excel/CSV file exported from the store's POS
(or maintained manually by staff) into catalog_items. This is the practical
"fetch stock from POS" mechanism until a real POS API is wired up via
catalog/pos_client.py.

Expected columns (header row, case-insensitive, any order):
    SKU | Name | Category | Unit | Price | Stock Qty

Usage:
    python -m catalog.excel_import path/to/file.xlsx
"""
import sys

from openpyxl import load_workbook

from catalog import store as catalog_store

REQUIRED_COLUMNS = {"name", "price", "stock_qty"}


def _normalize_key(header) -> str:
    return str(header).strip().lower().replace(" ", "_")


def import_file(path: str, source: str = "excel") -> int:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    headers = [_normalize_key(h) for h in rows[0]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise ValueError(
            f"Import file is missing required column(s): {', '.join(sorted(missing))}. "
            f"Expected headers like: SKU, Name, Category, Unit, Price, Stock Qty"
        )

    count = 0
    for raw_row in rows[1:]:
        if raw_row is None or all(v is None for v in raw_row):
            continue
        record = dict(zip(headers, raw_row))

        name = str(record.get("name") or "").strip()
        if not name:
            continue

        try:
            price = float(record.get("price") or 0)
        except (TypeError, ValueError):
            price = 0.0
        try:
            stock_qty = float(record.get("stock_qty") or 0)
        except (TypeError, ValueError):
            stock_qty = 0.0

        sku = str(record.get("sku")).strip() if record.get("sku") not in (None, "") else None
        category = str(record.get("category") or "").strip() or None
        unit = str(record.get("unit") or "").strip() or None

        catalog_store.upsert_item(
            name=name, price=price, stock_qty=stock_qty, sku=sku,
            category=category, unit=unit, source=source,
        )
        count += 1

    return count


def main():
    if len(sys.argv) < 2:
        print("Usage: python -m catalog.excel_import path/to/file.xlsx")
        sys.exit(1)
    path = sys.argv[1]
    count = import_file(path)
    print(f"Imported/updated {count} catalog item(s) from {path}")


if __name__ == "__main__":
    main()
